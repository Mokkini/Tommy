import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, timedelta
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import copy

# Konstanten
DATA_DIR = "monatsdaten"
COLUMNS = ["Datum", "Standort", "Disponent", "Fahrzeuge", "Stopps", 
           "Unverplante Stopps", "Kosten Fuhrpark", "Stoppschnitt", "Stoppkosten"]
KPI_COLUMNS = [
    ("Fahrzeuge", "Fahrzeuge"),
    ("Stopps", "Stopps"),
    ("Stoppschnitt", "Stoppschnitt"),
    ("Unverplante Stopps", "Unverplante Stopps"),
    ("Kosten Fuhrpark", "Kosten Fuhrpark"),
    ("Stoppkosten", "Stoppkosten")
]

# Session State für History
if 'history' not in st.session_state:
    st.session_state.history = []
if 'history_index' not in st.session_state:
    st.session_state.history_index = -1

def validate_data(df):
    """Validiert Daten und gibt Fehler zurück."""
    errors = []
    warnings = []
    
    for idx, row in df.iterrows():
        row_num = idx + 1
        
        # Datumsvalidierung
        if pd.notna(row['Datum']) and row['Datum'] != '':
            try:
                pd.to_datetime(row['Datum'], dayfirst=True)
            except:
                errors.append(f"Zeile {row_num}: Ungültiges Datum '{row['Datum']}' (Format: TT.MM.JJJJ)")
        
        # Numerische Felder validieren
        numeric_fields = ['Fahrzeuge', 'Stopps', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']
        for field in numeric_fields:
            if pd.notna(row[field]) and str(row[field]).strip() != '':
                try:
                    value = float(str(row[field]).replace(',', '.').replace('€', ''))
                    if value < 0:
                        errors.append(f"Zeile {row_num}: {field} darf nicht negativ sein ({value})")
                except:
                    errors.append(f"Zeile {row_num}: {field} ist keine gültige Zahl ('{row[field]}')")
        
        # Pflichtfelder prüfen
        if pd.isna(row['Datum']) or str(row['Datum']).strip() == '':
            warnings.append(f"Zeile {row_num}: Datum fehlt")
        if pd.isna(row['Standort']) or str(row['Standort']).strip() == '':
            warnings.append(f"Zeile {row_num}: Standort fehlt")
    
    return errors, warnings

def save_to_history(df, month_file):
    """Speichert aktuellen Zustand in der Historie."""
    if st.session_state.history_index < len(st.session_state.history) - 1:
        # Entferne alles nach dem aktuellen Index
        st.session_state.history = st.session_state.history[:st.session_state.history_index + 1]
    
    st.session_state.history.append({
        'df': df.copy(),
        'file': month_file,
        'timestamp': datetime.now()
    })
    
    # Limitiere Historie auf 20 Einträge
    if len(st.session_state.history) > 20:
        st.session_state.history.pop(0)
    
    st.session_state.history_index = len(st.session_state.history) - 1

def undo():
    """Macht die letzte Änderung rückgängig."""
    if st.session_state.history_index > 0:
        st.session_state.history_index -= 1
        return st.session_state.history[st.session_state.history_index]['df'].copy()
    return None

def redo():
    """Wiederholt die rückgängig gemachte Änderung."""
    if st.session_state.history_index < len(st.session_state.history) - 1:
        st.session_state.history_index += 1
        return st.session_state.history[st.session_state.history_index]['df'].copy()
    return None

def export_to_excel(df, month_name):
    """Exportiert DataFrame als formatiertes Excel mit Diagrammen."""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Daten"
    
    # Header-Style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header schreiben
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Daten schreiben
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Formatierung für bestimmte Spalten
            if df.columns[col_idx-1] == 'Stoppkosten':
                cell.number_format = '#,##0.00 €'
            elif df.columns[col_idx-1] in ['Fahrzeuge', 'Stopps', 'Unverplante Stopps']:
                cell.number_format = '#,##0'
            elif df.columns[col_idx-1] == 'Stoppschnitt':
                cell.number_format = '#,##0.0'
    
    # Spaltenbreite anpassen
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    # Chart erstellen (wenn numerische Daten vorhanden)
    if len(df) > 1:
        ws2 = wb.create_sheet(title="Diagramm")
        
        # Stoppkosten Chart
        chart = LineChart()
        chart.title = f"Stoppkosten Verlauf - {month_name}"
        chart.y_axis.title = 'Stoppkosten (€)'
        chart.x_axis.title = 'Datum'
        
        # Daten für Chart (nur wenn Stoppkosten-Spalte existiert)
        if 'Stoppkosten' in df.columns:
            data = Reference(ws, min_col=df.columns.get_loc('Stoppkosten')+1, 
                           min_row=1, max_row=len(df)+1)
            chart.add_data(data, titles_from_data=True)
            ws2.add_chart(chart, "A1")
    
    wb.save(output)
    output.seek(0)
    return output

def get_week_number(date):
    """Gibt ISO-Wochennummer zurück."""
    return date.isocalendar()[1]

def compare_weeks(df):
    """Vergleicht KPIs wochenweise."""
    df['Datum'] = pd.to_datetime(df['Datum'], dayfirst=True, errors='coerce')
    df = clean_kpi_columns(df)
    
    df['Woche'] = df['Datum'].apply(lambda x: f"KW {get_week_number(x)}" if pd.notna(x) else None)
    
    # WICHTIG: Erst pro Tag alle Standorte aggregieren, dann pro Woche
    # Tagessummen/-durchschnitte berechnen
    daily_data = df.groupby(['Datum', 'Woche']).agg({
        'Fahrzeuge': 'sum',  # Summe aller Standorte pro Tag
        'Stopps': 'sum',  # Summe aller Standorte pro Tag
        'Stoppschnitt': 'mean',  # Durchschnitt aller Standorte pro Tag
        'Unverplante Stopps': 'mean',  # Durchschnitt aller Standorte pro Tag
        'Kosten Fuhrpark': 'sum',  # Summe aller Standorte pro Tag
        'Stoppkosten': 'mean'  # Durchschnitt aller Standorte pro Tag
    }).reset_index()
    
    # Dann Wochendurchschnitte aus den Tageswerten berechnen
    weekly_data = daily_data.groupby('Woche').agg({
        'Fahrzeuge': 'mean',  # Ø Fahrzeuge pro Tag in der Woche
        'Stopps': 'mean',  # Ø Stopps pro Tag in der Woche
        'Stoppschnitt': 'mean',  # Ø Stoppschnitt der Woche
        'Unverplante Stopps': 'mean',  # Ø Unverplante Stopps der Woche
        'Kosten Fuhrpark': 'mean',  # Ø Kosten Fuhrpark der Woche
        'Stoppkosten': 'mean'  # Ø Stoppkosten der Woche
    }).reset_index()
    
    # Deltas berechnen
    for col in ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']:
        weekly_data[f'{col}_Delta'] = weekly_data[col].diff()
        pct_change = weekly_data[col].pct_change() * 100
        # Ersetze inf/-inf durch NaN (passiert bei Division durch 0)
        pct_change = pct_change.replace([float('inf'), float('-inf')], float('nan'))
        weekly_data[f'{col}_Delta%'] = pct_change.round(1)
    
    return weekly_data

def compare_months(month1_file, month2_file):
    """Vergleicht zwei Monate."""
    try:
        df1 = pd.read_csv(month1_file, sep=';')
        df2 = pd.read_csv(month2_file, sep=';')
        
        df1['Datum'] = pd.to_datetime(df1['Datum'], dayfirst=True, errors='coerce')
        df2['Datum'] = pd.to_datetime(df2['Datum'], dayfirst=True, errors='coerce')
        
        df1 = clean_kpi_columns(df1)
        df2 = clean_kpi_columns(df2)
        
        month1_name = os.path.basename(month1_file).replace('.csv', '')
        month2_name = os.path.basename(month2_file).replace('.csv', '')
        
        comparison = {
            'KPI': ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten'],
            month1_name: [
                df1['Fahrzeuge'].sum(),
                df1['Stopps'].sum(),
                df1['Stoppschnitt'].mean(),
                df1['Unverplante Stopps'].mean(),
                df1['Kosten Fuhrpark'].mean(),
                df1['Stoppkosten'].mean()
            ],
            month2_name: [
                df2['Fahrzeuge'].sum(),
                df2['Stopps'].sum(),
                df2['Stoppschnitt'].mean(),
                df2['Unverplante Stopps'].mean(),
                df2['Kosten Fuhrpark'].mean(),
                df2['Stoppkosten'].mean()
            ]
        }
        
        comp_df = pd.DataFrame(comparison)
        comp_df['Delta'] = comp_df[month2_name] - comp_df[month1_name]
        comp_df['Delta %'] = ((comp_df[month2_name] / comp_df[month1_name] - 1) * 100).round(1)
        
        return comp_df, df1, df2
    except Exception as e:
        st.error(f"Fehler beim Monatsvergleich: {e}")
        return None, None, None

def altair_chart_line(data, value_col, location_name):
    """Erstellt ein Altair Liniendiagramm für KPI-Verlauf."""
    return alt.Chart(data).mark_line(point=True).encode(
        x=alt.X('Datum:T', 
                title='Datum',
                axis=alt.Axis(format='%d.%m.%Y', formatType='time')),
        y=alt.Y(f'{value_col}:Q', 
                title=value_col,
                axis=alt.Axis(titleLimit=200)),
        tooltip=[
            alt.Tooltip('Datum:T', title='Datum', format='%d.%m.%Y'),
            alt.Tooltip(f'{value_col}:Q', title=value_col)
        ]
    ).properties(
        title=f"{location_name}",
        height=220
    ).configure_axis(
        labelFontSize=11,
        titleFontSize=12
    )

@st.cache_data
def load_master_standorte():
    try:
        master_csv = pd.read_csv("KPI_Eingabemaske.csv", sep=';', usecols=["Standort"])
        return sorted(master_csv['Standort'].dropna().unique().tolist())
    except (FileNotFoundError, pd.errors.EmptyDataError, KeyError):
        return []

def clean_kpi_columns(dataframe):
    """Bereinigt KPI-Spalten und konvertiert sie zu numerischen Werten."""
    kpi_cols = ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']
    for col in kpi_cols:
        if col in dataframe.columns:
            dataframe[col] = pd.to_numeric(
                dataframe[col]
                .astype(str)
                .str.replace('€', '', regex=False)
                .str.replace(',', '.', regex=False)
                .str.replace(' ', '', regex=False),
                errors='coerce'
            )
    return dataframe

# --- Navigation und Monatsverwaltung ---
st.set_page_config(page_title="KPI Dashboard", layout="wide")
st.sidebar.title("Navigation")
page = st.sidebar.radio("Gehe zu...", [
    "Eingabemaske",
    "Daily Report",
    "Wochenvergleich",
    "Monatsvergleich",
    "Verlauf (KPIs)"
])
st.sidebar.markdown("---")

# --- Monatsauswahl und CSV-Handling ---
os.makedirs(DATA_DIR, exist_ok=True)

def get_month_str(dt):
    return dt.strftime("%Y-%m")

def get_months_in_dir():
    files = [f for f in os.listdir(DATA_DIR) if f.endswith(".csv")]
    return sorted([f.replace(".csv", "") for f in files])

def get_month_file(month):
    return os.path.join(DATA_DIR, f"{month}.csv")

heute = datetime.today()
monate = get_months_in_dir()
aktueller_monat = get_month_str(heute)
if aktueller_monat not in monate:
    monate.append(aktueller_monat)
monate = sorted(list(set(monate)))

selected_month = st.sidebar.selectbox("Monat wählen", monate, index=monate.index(aktueller_monat) if aktueller_monat in monate else 0)

# Monat löschen (nur wenn leer)
st.sidebar.markdown("---")
if st.sidebar.button("🗑️ Ausgewählten Monat löschen", type="secondary", help="Nur unbefüllte Monate (ohne KPI-Daten) können gelöscht werden"):
    month_file = get_month_file(selected_month)
    if os.path.exists(month_file):
        try:
            df_check = pd.read_csv(month_file, sep=';')
            df_check = clean_kpi_columns(df_check)
            
            # Prüfe ob alle KPI-Spalten leer sind
            hat_daten = False
            for col in ['Fahrzeuge', 'Stopps', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']:
                if col in df_check.columns:
                    if df_check[col].notna().any() and (df_check[col] != '').any():
                        hat_daten = True
                        break
            
            if hat_daten:
                st.sidebar.error("❌ Monat hat Daten und kann nicht gelöscht werden!")
            else:
                os.remove(month_file)
                st.sidebar.success(f"✅ Monat {selected_month} gelöscht!")
                st.rerun()
        except Exception as e:
            st.sidebar.error(f"Fehler beim Löschen: {e}")
    else:
        st.sidebar.warning("Monat existiert nicht.")

# Neue Tabelle/Monat anlegen
st.sidebar.markdown("---")
new_month = st.sidebar.text_input("Neuer Monat (YYYY-MM)", value=get_month_str(heute), key="new_month_input")
if st.sidebar.button("Neuen Monat anlegen"):
    new_month_file = get_month_file(new_month)
    if new_month and not os.path.exists(new_month_file):
        # Erstelle Daten für alle Werktage des Monats mit allen Standorten
        try:
            year, month = map(int, new_month.split('-'))
            
            # Definiere Standorte basierend auf Monat
            if (year == 2025 and month >= 12) or year > 2025:
                # Ab Dezember 2025 und alle zukünftigen Monate: Produktionsstandorte
                standorte = [
                    'Delmenhorst',
                    'Güstrow',
                    'Döbeln',
                    'Melle',
                    'Langenfeld',
                    'Kassel',
                    'Berlin',
                    'Aschaffenburg',
                    'Renningen'
                ]
            else:
                # Nur Oktober/November 2025: Alte Standorte
                standorte = [
                    'Aschaffenburg',
                    'Renningen',
                    'Hamburg',
                    'Hannover',
                    'Langenfeld',
                    'Föhren',
                    'Kassel',
                    'Stockstadt',
                    'Eutingen',
                    'Berlin',
                    'Melle',
                    'Delmenhorst',
                    'Güstrow'
                ]
            
            # Generiere alle Werktage des Monats
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            
            dates = pd.date_range(start=start_date, end=end_date, freq='D')
            werktage = [d for d in dates if d.weekday() < 5]  # Mo-Fr
            
            # Erstelle Einträge für jeden Werktag und jeden Standort
            data = []
            for date in werktage:
                for standort in standorte:
                    data.append({
                        'Datum': date.strftime('%d.%m.%Y'),
                        'Standort': standort,
                        'Disponent': '',
                        'Fahrzeuge': '',
                        'Stopps': '',
                        'Unverplante Stopps': '',
                        'Kosten Fuhrpark': '',
                        'Stoppschnitt': '',
                        'Stoppkosten': ''
                    })
            
            # Speichere als CSV
            df_new = pd.DataFrame(data)
            df_new.to_csv(get_month_file(new_month), sep=';', index=False)
            st.success(f"✅ Monat {new_month} angelegt mit {len(werktage)} Werktagen und {len(standorte)} Standorten!")
            st.rerun()
        except Exception as e:
            st.error(f"Fehler beim Erstellen des Monats: {e}")
    else:
        st.warning("Monat existiert bereits oder das Format ist ungültig.")

# Daten für gewählten Monat laden
month_file = get_month_file(selected_month)
if os.path.exists(month_file):
    try:
        df = pd.read_csv(month_file, sep=';')
    except (pd.errors.EmptyDataError, pd.errors.ParserError):
        st.warning(f"Fehler beim Laden von {selected_month}. Neue Tabelle wird erstellt.")
        df = pd.DataFrame(columns=COLUMNS)
else:
    df = pd.DataFrame(columns=COLUMNS)

# Seitenlogik
if page == "Eingabemaske":
    st.header(f"📋 Eingabemaske ({selected_month})")
    
    # Undo/Redo Buttons
    col1, col2, col3, col4 = st.columns([1, 1, 2, 2])
    with col1:
        if st.button("↶ Undo", disabled=st.session_state.history_index <= 0, 
                    help="Letzte Änderung rückgängig machen"):
            restored_df = undo()
            if restored_df is not None:
                df = restored_df
                df.to_csv(month_file, sep=';', index=False)
                st.success("✅ Änderung rückgängig gemacht!")
                st.rerun()
    
    with col2:
        if st.button("↷ Redo", disabled=st.session_state.history_index >= len(st.session_state.history) - 1,
                    help="Rückgängig gemachte Änderung wiederholen"):
            restored_df = redo()
            if restored_df is not None:
                df = restored_df
                df.to_csv(month_file, sep=';', index=False)
                st.success("✅ Änderung wiederhergestellt!")
                st.rerun()
    
    with col3:
        st.info(f"📊 {len(df)} Einträge | History: {st.session_state.history_index + 1}/{len(st.session_state.history)}")
    
    st.write("Bearbeite die Tabelle direkt im Editor unten.")
    
    # Bereinige DataFrame für Editor
    df_display = df.copy()
    for col in df_display.columns:
        df_display[col] = df_display[col].astype(str).replace('nan', '')
        if col in ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']:
            df_display[col] = df_display[col].apply(lambda x: x.rstrip('0').rstrip('.') if '.' in str(x) and str(x) not in ['', 'nan'] else x)
    
    # Konfiguration für die Spalten
    column_config = {
        "Datum": st.column_config.TextColumn("Datum", help="Format: TT.MM.JJJJ", required=True),
        "Standort": st.column_config.TextColumn("Standort", width="medium", required=True),
        "Disponent": st.column_config.TextColumn("Disponent", width="medium"),
        "Fahrzeuge": st.column_config.TextColumn("Fahrzeuge", help="Anzahl eingesetzter Fahrzeuge"),
        "Stopps": st.column_config.TextColumn("Stopps", help="Anzahl Stopps gesamt"),
        "Unverplante Stopps": st.column_config.TextColumn("Unverplante Stopps"),
        "Kosten Fuhrpark": st.column_config.TextColumn("Kosten Fuhrpark", help="Gesamtkosten Fuhrpark in Euro"),
        "Stoppschnitt": st.column_config.TextColumn("Stoppschnitt", 
                                                     help="Wird automatisch berechnet: Stopps ÷ Fahrzeuge",
                                                     disabled=True),
        "Stoppkosten": st.column_config.TextColumn("Stoppkosten", 
                                                     help="Wird automatisch berechnet: Kosten Fuhrpark ÷ Stopps",
                                                     disabled=True),
    }
    
    edited_df = st.data_editor(
        df_display, 
        num_rows="dynamic", 
        width='stretch', 
        key=f"editor_{selected_month}",
        column_config=column_config,
        hide_index=True
    )
    
    # Berechne Stoppschnitt und Stoppkosten automatisch
    if not edited_df.empty:
        for idx in edited_df.index:
            fahrzeuge_str = str(edited_df.at[idx, 'Fahrzeuge']).strip()
            stopps_str = str(edited_df.at[idx, 'Stopps']).strip()
            kosten_fuhrpark_str = str(edited_df.at[idx, 'Kosten Fuhrpark']).strip()
            
            fahrzeuge = pd.to_numeric(fahrzeuge_str, errors='coerce') if fahrzeuge_str and fahrzeuge_str != '' else None
            stopps = pd.to_numeric(stopps_str, errors='coerce') if stopps_str and stopps_str != '' else None
            kosten_fuhrpark = pd.to_numeric(kosten_fuhrpark_str.replace(',', '.'), errors='coerce') if kosten_fuhrpark_str and kosten_fuhrpark_str != '' else None
            
            # Berechne Stoppschnitt
            if pd.notna(fahrzeuge) and pd.notna(stopps) and fahrzeuge > 0:
                stoppschnitt = round(stopps / fahrzeuge, 1)
                edited_df.at[idx, 'Stoppschnitt'] = str(stoppschnitt).replace('.', ',')
            else:
                edited_df.at[idx, 'Stoppschnitt'] = ''
            
            # Berechne Stoppkosten
            if pd.notna(kosten_fuhrpark) and pd.notna(stopps) and stopps > 0:
                stoppkosten = round(kosten_fuhrpark / stopps, 2)
                edited_df.at[idx, 'Stoppkosten'] = str(stoppkosten).replace('.', ',')
            else:
                edited_df.at[idx, 'Stoppkosten'] = ''
    
    # Validierung
    errors, warnings = validate_data(edited_df)
    
    if errors:
        st.error("🚨 **Validierungsfehler:**")
        for error in errors:
            st.error(f"❌ {error}")
    
    if warnings:
        st.warning("⚠️ **Warnungen:**")
        for warning in warnings:
            st.warning(f"⚠️ {warning}")
    
    st.info("💡 Tipp: Stoppschnitt und Stoppkosten werden automatisch berechnet. Klicke auf 'Tabelle speichern' um Änderungen zu sichern.")
    
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        csv = edited_df.to_csv(index=False, sep=';').encode('utf-8')
        st.download_button("📥 CSV Export", csv, f"KPI_Eingabemaske_{selected_month}.csv", "text/csv")
    
    with col2:
        excel_data = export_to_excel(edited_df, selected_month)
        st.download_button("📊 Excel Export", excel_data, 
                          f"KPI_Dashboard_{selected_month}.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    with col3:
        if st.button("💾 Tabelle speichern", key=f"save_{selected_month}", type="primary", disabled=len(errors) > 0):
            save_to_history(df, month_file)
            edited_df.to_csv(month_file, sep=';', index=False)
            st.success("✅ Tabelle gespeichert!")
            st.rerun()
    
    with col4:
        if st.button("🔄 Neu laden", key=f"reload_{selected_month}"):
            st.rerun()

elif page == "Daily Report":
    st.header("📊 Daily Report")
    if df is not None and not df.empty:
        df['Datum'] = pd.to_datetime(df['Datum'], dayfirst=True, errors='coerce')
        df = clean_kpi_columns(df)

        # Filtere nur Zeilen mit tatsächlichen Daten (mindestens Stopps oder Fahrzeuge vorhanden)
        df_with_data = df[
            (df['Stopps'].notna() & (df['Stopps'] > 0)) | 
            (df['Fahrzeuge'].notna() & (df['Fahrzeuge'] > 0))
        ]
        
        if df_with_data.empty:
            st.info("Keine Daten mit KPI-Werten vorhanden.")
        else:
            latest_date = df_with_data['Datum'].max()
            if pd.isna(latest_date):
                st.info("Keine gültigen Datumswerte vorhanden.")
            else:
                df_latest = df_with_data[df_with_data['Datum'] == latest_date]
            
            # Finde den Vortag (letztes verfügbares Datum vor dem aktuellen - nur mit Daten)
            dates = sorted(df_with_data['Datum'].dropna().unique())
            previous_date = dates[-2] if len(dates) >= 2 else None
            df_previous = df_with_data[df_with_data['Datum'] == previous_date] if previous_date else pd.DataFrame()

            kpi_stoppkosten = df_latest['Stoppkosten'].mean()
            kpi_stopps = df_latest['Stopps'].sum()
            kpi_fahrzeuge = df_latest['Fahrzeuge'].sum()
            kpi_stoppschnitt = df_latest['Stoppschnitt'].mean()
            kpi_unverplant = df_latest['Unverplante Stopps'].mean()
            
            # Berechne Deltas zum Vortag
            delta_stoppkosten = None
            delta_stopps = None
            delta_fahrzeuge = None
            if not df_previous.empty:
                delta_stoppkosten = kpi_stoppkosten - df_previous['Stoppkosten'].mean()
                delta_stopps = kpi_stopps - df_previous['Stopps'].sum()
                delta_fahrzeuge = kpi_fahrzeuge - df_previous['Fahrzeuge'].sum()

            # Zeige Datum mit Vortag-Info
            vortag_info = f" (Vergleich zu {previous_date.strftime('%d.%m.%Y')})" if previous_date else ""
            st.subheader(f"📅 Letzter Tag mit Daten: {latest_date.strftime('%d.%m.%Y')}{vortag_info}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Ø Stoppkosten", f"{kpi_stoppkosten:.2f} €",
                       delta=f"{delta_stoppkosten:.2f} €" if delta_stoppkosten is not None else None,
                       delta_color="inverse")
            col2.metric("Gesamt Stopps", f"{int(kpi_stopps):,}",
                       delta=f"{int(delta_stopps):,}" if delta_stopps is not None else None)
            col3.metric("Fahrzeuge", f"{int(kpi_fahrzeuge)}",
                       delta=f"{int(delta_fahrzeuge)}" if delta_fahrzeuge is not None else None)
            col4.metric("Ø Stoppschnitt", f"{kpi_stoppschnitt:.1f}")
            
            st.markdown("---")
            
            # Excel Export für Daily Report
            col1, col2 = st.columns([4, 1])
            with col2:
                excel_report = export_to_excel(df_latest, f"Daily_Report_{latest_date.strftime('%d.%m.%Y')}")
                st.download_button("📊 Report als Excel", excel_report,
                                 f"Daily_Report_{latest_date.strftime('%d%m%Y')}.xlsx",
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            # Rest vom Daily Report (gekürzt für Platzgründe - wie vorher)
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Ø Unverplante Stopps", f"{kpi_unverplant:.1f}")
            with col2:
                effizienz = (kpi_stopps / kpi_fahrzeuge) if kpi_fahrzeuge > 0 else 0
                st.metric("Effizienz (Stopps/Fzg)", f"{effizienz:.1f}")
            with col3:
                aktive_standorte = len(df_latest)
                st.metric("Aktive Standorte", f"{aktive_standorte}")
            
            st.markdown("---")
            st.subheader("🏆 Standort-Performance")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**🟢 TOP 5 - Niedrigste Stoppkosten**")
                top5 = df_latest.nsmallest(5, 'Stoppkosten')[['Standort', 'Stoppkosten', 'Stopps', 'Fahrzeuge']]
                top5['Stoppkosten'] = top5['Stoppkosten'].apply(lambda x: f"{x:.2f} €")
                st.dataframe(top5, hide_index=True, width='stretch')
            
            with col2:
                st.markdown("**🔴 BOTTOM 5 - Höchste Stoppkosten**")
                bottom5 = df_latest.nlargest(5, 'Stoppkosten')[['Standort', 'Stoppkosten', 'Stopps', 'Fahrzeuge']]
                bottom5['Stoppkosten'] = bottom5['Stoppkosten'].apply(lambda x: f"{x:.2f} €")
                st.dataframe(bottom5, hide_index=True, width='stretch')
            
            st.markdown("---")
            st.info(f"💡 Hinweis: Es werden nur Tage mit ausgefüllten KPI-Daten angezeigt. Aktuell: {latest_date.strftime('%d.%m.%Y')}")
    else:
        st.info("Keine Daten vorhanden.")

elif page == "Wochenvergleich":
    st.header("📊 Wochenvergleich")
    
    if df is not None and not df.empty:
        weekly_data = compare_weeks(df)
        
        if not weekly_data.empty:
            st.subheader(f"KPIs pro Woche - {selected_month}")
            
            # Formatiere Tabelle für bessere Lesbarkeit
            display_data = weekly_data.copy()
            
            # Runde numerische Werte
            for col in display_data.columns:
                if 'Delta%' in col:
                    display_data[col] = display_data[col].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "N/A")
                    # Entferne .0 bei ganzen Zahlen (z.B. +100.0% -> +100%)
                    display_data[col] = display_data[col].str.replace(r'\.0%$', '%', regex=True)
                elif 'Delta' in col and col != 'Woche':
                    display_data[col] = display_data[col].apply(lambda x: f"{x:+.1f}" if pd.notna(x) else "")
                    # Entferne .0 bei ganzen Zahlen (z.B. +501.0 -> +501)
                    display_data[col] = display_data[col].str.replace(r'\.0$', '', regex=True)
                elif col in ['Stoppkosten', 'Kosten Fuhrpark']:
                    display_data[col] = display_data[col].apply(lambda x: f"{x:.2f} €" if pd.notna(x) else "")
                elif col in ['Stoppschnitt', 'Unverplante Stopps']:
                    display_data[col] = display_data[col].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "")
                    # Entferne .0 bei ganzen Zahlen
                    display_data[col] = display_data[col].str.replace(r'\.0$', '', regex=True)
                elif col in ['Fahrzeuge', 'Stopps']:
                    display_data[col] = display_data[col].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")
            
            # Benenne Spalten um für bessere Lesbarkeit
            display_data = display_data.rename(columns={
                'Fahrzeuge': 'Ø Fzg/Tag',
                'Fahrzeuge_Delta': 'Fzg Δ',
                'Fahrzeuge_Delta%': 'Fzg Δ%',
                'Stopps': 'Ø Stopps/Tag',
                'Stopps_Delta': 'Stopps Δ',
                'Stopps_Delta%': 'Stopps Δ%',
                'Stoppschnitt': 'Ø Schnitt',
                'Stoppschnitt_Delta': 'Schnitt Δ',
                'Stoppschnitt_Delta%': 'Schnitt Δ%',
                'Unverplante Stopps': 'Ø Unverplant',
                'Unverplante Stopps_Delta': 'Unverplant Δ',
                'Unverplante Stopps_Delta%': 'Unverplant Δ%',
                'Kosten Fuhrpark': 'Ø Fuhrpark',
                'Kosten Fuhrpark_Delta': 'Fuhrpark Δ',
                'Kosten Fuhrpark_Delta%': 'Fuhrpark Δ%',
                'Stoppkosten': 'Ø Stoppkosten',
                'Stoppkosten_Delta': 'Stoppkosten Δ',
                'Stoppkosten_Delta%': 'Stoppkosten Δ%'
            })
            
            st.dataframe(display_data, width='stretch', hide_index=True)
            
            # Visualisierung
            st.markdown("---")
            st.subheader("📈 Wochentrends")
            
            col1, col2 = st.columns(2)
            
            with col1:
                chart_stopps = alt.Chart(weekly_data).mark_line(point=True, color='#4ECDC4').encode(
                    x=alt.X('Woche:N', title='Woche'),
                    y=alt.Y('Stopps:Q', title='Stopps'),
                    tooltip=['Woche', 'Stopps']
                ).properties(title='Stopps pro Woche', height=300)
                st.altair_chart(chart_stopps, width='stretch')
            
            with col2:
                chart_kosten = alt.Chart(weekly_data).mark_line(point=True, color='#FF6B6B').encode(
                    x=alt.X('Woche:N', title='Woche'),
                    y=alt.Y('Stoppkosten:Q', title='Stoppkosten (€)'),
                    tooltip=['Woche', alt.Tooltip('Stoppkosten:Q', format='.2f')]
                ).properties(title='Stoppkosten pro Woche', height=300)
                st.altair_chart(chart_kosten, width='stretch')
            
            # Excel Export
            excel_data = export_to_excel(weekly_data, f"Wochenvergleich_{selected_month}")
            st.download_button("📊 Wochenvergleich als Excel", excel_data,
                             f"Wochenvergleich_{selected_month}.xlsx",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Nicht genügend Daten für Wochenvergleich.")
    else:
        st.info("Keine Daten vorhanden.")

elif page == "Monatsvergleich":
    st.header("📊 Monatsvergleich")
    
    # Modus-Auswahl
    vergleich_modus = st.radio("Vergleichsmodus", 
                               ["2 Monate vergleichen", "Mehrere Monate analysieren"],
                               horizontal=True)
    
    if vergleich_modus == "2 Monate vergleichen":
        # Original 2-Monats-Vergleich
        if len(monate) >= 2:
            col1, col2 = st.columns(2)
            with col1:
                month1 = st.selectbox("Monat 1", monate, index=len(monate)-2 if len(monate) >= 2 else 0)
            with col2:
                month2 = st.selectbox("Monat 2", monate, index=len(monate)-1)
            
            if st.button("🔄 Monate vergleichen", type="primary"):
                month1_file = get_month_file(month1)
                month2_file = get_month_file(month2)
                
                comp_df, df1, df2 = compare_months(month1_file, month2_file)
                
                if comp_df is not None:
                    st.subheader(f"Vergleich: {month1} vs {month2}")
                    
                    # Formatiere Ausgabe
                    comp_display = comp_df.copy()
                    for col in [month1, month2, 'Delta']:
                        comp_display[col] = comp_display[col].round(2)
                    
                    st.dataframe(comp_display, width='stretch', hide_index=True)
                    
                    # Visualisierung
                    st.markdown("---")
                st.subheader("📈 Monatsvergleich Visualisierung")
                
                # Umformen für Altair
                comparison_long = comp_df.melt(id_vars=['KPI'], value_vars=[month1, month2],
                                              var_name='Monat', value_name='Wert')
                
                chart = alt.Chart(comparison_long).mark_bar().encode(
                    x=alt.X('KPI:N', title='KPI'),
                    y=alt.Y('Wert:Q', title='Wert'),
                    color=alt.Color('Monat:N', scale=alt.Scale(scheme='category10')),
                    xOffset='Monat:N',
                    tooltip=['KPI', 'Monat', alt.Tooltip('Wert:Q', format='.2f')]
                ).properties(height=400)
                
                st.altair_chart(chart, use_container_width=True)
                
                # Excel Export
                excel_data = export_to_excel(comp_df, f"Vergleich_{month1}_vs_{month2}")
                st.download_button("📊 Vergleich als Excel", excel_data,
                                 f"Monatsvergleich_{month1}_vs_{month2}.xlsx",
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Mindestens 2 Monate benötigt für Vergleich.")
    
    else:  # Mehrere Monate analysieren
        st.subheader("📊 Multi-Monats-Analyse")
        
        if len(monate) >= 2:
            # Monat-Auswahl mit Multiselect
            selected_months = st.multiselect(
                "Monate auswählen (chronologisch sortiert)",
                monate,
                default=monate[-3:] if len(monate) >= 3 else monate
            )
            
            if len(selected_months) < 2:
                st.warning("⚠️ Bitte mindestens 2 Monate auswählen.")
            else:
                if st.button("📊 Analyse starten", type="primary"):
                    # Lade Daten für alle ausgewählten Monate
                    all_data = {}
                    for month in selected_months:
                        try:
                            month_file = get_month_file(month)
                            df_month = pd.read_csv(month_file, sep=';')
                            df_month['Datum'] = pd.to_datetime(df_month['Datum'], dayfirst=True, errors='coerce')
                            df_month = clean_kpi_columns(df_month)
                            all_data[month] = df_month
                        except Exception as e:
                            st.error(f"Fehler beim Laden von {month}: {e}")
                    
                    if len(all_data) >= 2:
                        # Berechne KPIs für jeden Monat
                        kpi_comparison = {
                            'Monat': [],
                            'Fahrzeuge': [],
                            'Stopps': [],
                            'Stoppschnitt': [],
                            'Unverplante Stopps': [],
                            'Kosten Fuhrpark': [],
                            'Stoppkosten': []
                        }
                        
                        for month in selected_months:
                            if month in all_data:
                                df_m = all_data[month]
                                kpi_comparison['Monat'].append(month)
                                kpi_comparison['Fahrzeuge'].append(df_m['Fahrzeuge'].sum())
                                kpi_comparison['Stopps'].append(df_m['Stopps'].sum())
                                kpi_comparison['Stoppschnitt'].append(df_m['Stoppschnitt'].mean())
                                kpi_comparison['Unverplante Stopps'].append(df_m['Unverplante Stopps'].mean())
                                kpi_comparison['Kosten Fuhrpark'].append(df_m['Kosten Fuhrpark'].mean())
                                kpi_comparison['Stoppkosten'].append(df_m['Stoppkosten'].mean())
                        
                        multi_df = pd.DataFrame(kpi_comparison)
                        
                        # Zeige Tabelle mit allen Monaten
                        st.subheader(f"📋 Übersicht: {len(selected_months)} Monate")
                        
                        # Formatiere für Anzeige
                        display_df = multi_df.copy()
                        for col in ['Fahrzeuge', 'Stopps']:
                            display_df[col] = display_df[col].astype(int)
                        for col in ['Stoppschnitt', 'Unverplante Stopps', 'Stoppkosten']:
                            display_df[col] = display_df[col].round(2)
                        
                        st.dataframe(display_df, width='stretch', hide_index=True)
                        
                        # Trend-Analyse
                        st.markdown("---")
                        st.subheader("📈 Trend-Analyse über ausgewählte Monate")
                        
                        # KPI-Auswahl für Diagramm
                        kpi_select = st.selectbox(
                            "KPI für Trendanalyse",
                            ['Stopps', 'Stoppkosten', 'Kosten Fuhrpark', 'Stoppschnitt', 'Fahrzeuge', 'Unverplante Stopps'],
                            index=1
                        )
                        
                        # Liniendiagramm
                        trend_chart = alt.Chart(multi_df).mark_line(point=True, size=3).encode(
                            x=alt.X('Monat:N', title='Monat', sort=selected_months),
                            y=alt.Y(f'{kpi_select}:Q', title=kpi_select),
                            tooltip=['Monat', alt.Tooltip(f'{kpi_select}:Q', format='.2f')]
                        ).properties(
                            height=400,
                            title=f'{kpi_select} Verlauf über {len(selected_months)} Monate'
                        )
                        
                        st.altair_chart(trend_chart, use_container_width=True)
                        
                        # Vergleich mit vorherigem Monat (Delta)
                        st.markdown("---")
                        st.subheader("📊 Monat-zu-Monat Veränderungen")
                        
                        delta_df = multi_df.copy()
                        for col in ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']:
                            delta_df[f'{col}_Delta'] = delta_df[col].diff()
                            pct_change = delta_df[col].pct_change() * 100
                            # Ersetze inf/-inf durch NaN
                            pct_change = pct_change.replace([float('inf'), float('-inf')], float('nan'))
                            delta_df[f'{col}_Delta%'] = pct_change.round(1)
                        
                        # Zeige Delta-Tabelle
                        delta_display = delta_df[['Monat', 'Stoppkosten', 'Stoppkosten_Delta', 'Stoppkosten_Delta%',
                                                   'Stopps', 'Stopps_Delta', 'Fahrzeuge', 'Fahrzeuge_Delta']].copy()
                        
                        for col in ['Stoppkosten', 'Stoppkosten_Delta']:
                            delta_display[col] = delta_display[col].round(2)
                        
                        st.dataframe(delta_display, width='stretch', hide_index=True)
                        
                        # Heatmap für alle KPIs
                        st.markdown("---")
                        st.subheader("🔥 KPI Heatmap")
                        
                        # Normalisiere Daten für Heatmap (0-100 Scale)
                        heatmap_data = []
                        kpi_cols = ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']
                        
                        for kpi in kpi_cols:
                            for idx, month in enumerate(multi_df['Monat']):
                                value = multi_df.loc[idx, kpi]
                                heatmap_data.append({
                                    'Monat': month,
                                    'KPI': kpi,
                                    'Wert': value
                                })
                        
                        heatmap_df = pd.DataFrame(heatmap_data)
                        
                        heatmap_chart = alt.Chart(heatmap_df).mark_rect().encode(
                            x=alt.X('Monat:N', title='Monat', sort=selected_months),
                            y=alt.Y('KPI:N', title='KPI'),
                            color=alt.Color('Wert:Q', scale=alt.Scale(scheme='blues')),
                            tooltip=['Monat', 'KPI', alt.Tooltip('Wert:Q', format='.2f')]
                        ).properties(
                            height=300,
                            title='KPI Heatmap über alle Monate'
                        )
                        
                        st.altair_chart(heatmap_chart, use_container_width=True)
                        
                        # Excel Export
                        excel_data = export_to_excel(display_df, f"Multi_Monats_Analyse")
                        st.download_button("📊 Multi-Monats-Analyse als Excel", excel_data,
                                         f"Multi_Monats_Analyse_{len(selected_months)}_Monate.xlsx",
                                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Mindestens 2 Monate benötigt für Multi-Monats-Analyse.")

elif page == "Verlauf (KPIs)":
    st.header("📈 Verlauf (KPIs)")
    if df is not None and not df.empty:
        df['Datum'] = pd.to_datetime(df['Datum'], dayfirst=True, errors='coerce')
        df = clean_kpi_columns(df)

        standorte = sorted(df['Standort'].dropna().unique())
        selected_standorte = st.multiselect(
            "Standorte auswählen", 
            standorte, 
            default=standorte[:1] if standorte else []
        )
        if not selected_standorte:
            st.info("Bitte mindestens einen Standort auswählen.")
        else:
            for selected_standort in selected_standorte:
                st.subheader(f"📍 {selected_standort}")
                for kpi_name, kpi_col in KPI_COLUMNS:
                    chart_data = df[df['Standort'] == selected_standort][['Datum', kpi_col]].sort_values('Datum')
                    if chart_data.empty:
                        st.info(f"Keine Daten für {selected_standort} / {kpi_name}.")
                    else:
                        chart = altair_chart_line(chart_data, kpi_col, kpi_name)
                        st.altair_chart(chart, width='stretch')
    else:
        st.info("Keine Daten vorhanden.")
