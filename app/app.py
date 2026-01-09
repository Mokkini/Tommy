"""
KPI Dashboard - Streamlit App mit SQLite und Login
Für Deployment auf Synology NAS via Docker
"""
import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, timedelta
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Lokale Imports
from database import (
    init_database, get_months, get_month_data, save_month_data,
    create_month, delete_month, get_standorte, init_default_standorte
)
from auth import require_auth, show_user_info, is_admin

# Initialisiere Datenbank beim Start
init_database()
init_default_standorte()

# Konstanten
COLUMNS = ["Datum", "Standort", "Disponent", "Fahrzeuge", "Stopps", 
           "Unverplante Stopps", "Kosten Fuhrpark", "Stoppschnitt", "Stoppkosten"]
KPI_COLUMNS = [
    ("Fahrzeuge", "Fahrzeuge"),
    ("Stopps", "Stopps"),
    ("Stoppschnitt", "Stoppschnitt"),
    ("Unverplante Stopps", "Unverplante Stopps"),
    ("Kosten Fuhrpark", "Kosten Fuhrpark"),
    ("Stoppkosten", "Stoppkosten")
]

# Standard-Standorte
DEFAULT_STANDORTE = [
    'Delmenhorst', 'Güstrow', 'Döbeln', 'Melle', 'Langenfeld',
    'Kassel', 'Berlin', 'Aschaffenburg', 'Renningen'
]

# --- Page Config ---
st.set_page_config(
    page_title="KPI Dashboard - Dispo",
    page_icon="📊",
    layout="wide",
    menu_items={
        "Get help": None,
        "Report a bug": None,
        "About": None
    }
)

# CSS: Versteckt Streamlit-Menü, Header, Footer und Toolbar-Icons
hide_streamlit_style = """
<style>
#MainMenu {visibility: hidden;}
header {visibility: hidden;}
footer {visibility: hidden;}
.stAppDeployButton {display: none;}
.stAppViewBlockContainer > div:first-child {display: none;}
button[title="View app source"] {display: none;}
button[title="Manage app"] {display: none;}
a[href*="github"] {display: none;}
.viewerBadge_container__1QSob {display: none;}
.styles_viewerBadge__1yB5_ {display: none;}
[data-testid="stToolbar"] {display: none;}
[data-testid="manage-app-button"] {display: none;}
</style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# --- Authentifizierung ---
require_auth()

# --- Hilfsfunktionen ---
def clean_kpi_columns(dataframe):
    """Bereinigt KPI-Spalten und konvertiert sie zu numerischen Werten."""
    kpi_cols = ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']
    for col in kpi_cols:
        if col in dataframe.columns:
            dataframe[col] = pd.to_numeric(
                dataframe[col]
                .astype(str)
                .str.replace('€', '', regex=False)
                .str.replace(',', '.', regex=False)
                .str.replace(' ', '', regex=False),
                errors='coerce'
            )
    return dataframe

def validate_data(df):
    """Validiert Daten und gibt Fehler zurück."""
    errors = []
    warnings = []
    
    for idx, row in df.iterrows():
        row_num = idx + 1
        
        # Datumsvalidierung
        if pd.notna(row['Datum']) and row['Datum'] != '':
            try:
                pd.to_datetime(row['Datum'], dayfirst=True)
            except:
                errors.append(f"Zeile {row_num}: Ungültiges Datum '{row['Datum']}'")
        
        # Numerische Felder validieren
        numeric_fields = ['Fahrzeuge', 'Stopps', 'Unverplante Stopps', 'Kosten Fuhrpark']
        for field in numeric_fields:
            if pd.notna(row.get(field)) and str(row.get(field, '')).strip() != '':
                try:
                    value = float(str(row[field]).replace(',', '.').replace('€', ''))
                    if value < 0:
                        errors.append(f"Zeile {row_num}: {field} darf nicht negativ sein")
                except:
                    errors.append(f"Zeile {row_num}: {field} ist keine gültige Zahl")
    
    return errors, warnings

def export_to_excel(df, title):
    """Exportiert DataFrame als formatiertes Excel."""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Daten"
    
    # Header-Style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header schreiben
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Daten schreiben
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Spaltenbreite anpassen
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    wb.save(output)
    output.seek(0)
    return output

def get_week_number(date):
    """Gibt ISO-Wochennummer zurück."""
    return date.isocalendar()[1]

def format_number_de(value, decimals=2):
    """Formatiert Zahlen im deutschen Format (Komma als Dezimaltrennzeichen)."""
    if pd.isna(value) or value == '':
        return ''
    try:
        num = float(value)
        if decimals == 0:
            return f"{int(num):,}".replace(',', '.')
        else:
            formatted = f"{num:.{decimals}f}".replace('.', ',')
            # Tausendertrennzeichen
            parts = formatted.split(',')
            parts[0] = f"{int(parts[0]):,}".replace(',', '.')
            return ','.join(parts)
    except:
        return str(value)

def compare_weeks(df):
    """Vergleicht KPIs wochenweise."""
    df = df.copy()
    df['Datum'] = pd.to_datetime(df['Datum'], dayfirst=True, errors='coerce')
    df = clean_kpi_columns(df)
    
    df['Woche'] = df['Datum'].apply(lambda x: f"KW {get_week_number(x)}" if pd.notna(x) else None)
    
    # Tagessummen berechnen
    daily_data = df.groupby(['Datum', 'Woche']).agg({
        'Fahrzeuge': 'sum',
        'Stopps': 'sum',
        'Stoppschnitt': 'mean',
        'Unverplante Stopps': 'mean',
        'Kosten Fuhrpark': 'sum',
        'Stoppkosten': 'mean'
    }).reset_index()
    
    # Wochendurchschnitte
    weekly_data = daily_data.groupby('Woche').agg({
        'Fahrzeuge': 'mean',
        'Stopps': 'mean',
        'Stoppschnitt': 'mean',
        'Unverplante Stopps': 'mean',
        'Kosten Fuhrpark': 'mean',
        'Stoppkosten': 'mean'
    }).reset_index()
    
    # Deltas berechnen
    for col in ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppkosten']:
        weekly_data[f'{col}_Delta'] = weekly_data[col].diff()
        pct_change = weekly_data[col].pct_change() * 100
        pct_change = pct_change.replace([float('inf'), float('-inf')], float('nan'))
        weekly_data[f'{col}_Delta%'] = pct_change.round(1)
    
    return weekly_data

# --- Navigation (rollenbasiert) ---
st.sidebar.title("📊 KPI Dashboard")
show_user_info()

# Disponenten sehen nur die Eingabemaske
if is_admin():
    # Admin sieht alle Seiten
    page = st.sidebar.radio("Navigation", [
        "📋 Eingabemaske",
        "📊 Daily Report",
        "📈 Wochenvergleich",
        "📅 Monatsvergleich",
        "📉 Verlauf (KPIs)"
    ])
else:
    # Disponent sieht nur Eingabemaske
    page = "📋 Eingabemaske"
    st.sidebar.info("📝 Sie haben Zugriff auf die Eingabemaske.")

st.sidebar.markdown("---")

# --- Monatsauswahl ---
def get_month_str(dt):
    return dt.strftime("%Y-%m")

heute = datetime.today()
monate = get_months()
aktueller_monat = get_month_str(heute)

if aktueller_monat not in monate:
    monate.append(aktueller_monat)
monate = sorted(list(set(monate)))

if not monate:
    monate = [aktueller_monat]

selected_month = st.sidebar.selectbox(
    "Monat wählen", 
    monate, 
    index=monate.index(aktueller_monat) if aktueller_monat in monate else 0
)

# Neuen Monat anlegen
st.sidebar.markdown("---")
new_month = st.sidebar.text_input("Neuer Monat (YYYY-MM)", value=get_month_str(heute))
if st.sidebar.button("➕ Monat anlegen"):
    if new_month and new_month not in monate:
        try:
            create_month(new_month, DEFAULT_STANDORTE)
            st.sidebar.success(f"✅ Monat {new_month} angelegt!")
            st.rerun()
        except Exception as e:
            st.sidebar.error(f"Fehler: {e}")

# --- Daten laden ---
df = get_month_data(selected_month)
if df.empty:
    df = pd.DataFrame(columns=COLUMNS)

# === SEITEN ===

if page == "📋 Eingabemaske":
    st.header(f"📋 Eingabemaske ({selected_month})")
    
    st.write("Bearbeite die Tabelle direkt im Editor:")
    
    # Bereinige für Anzeige
    df_display = df.copy()
    for col in df_display.columns:
        df_display[col] = df_display[col].astype(str).replace('nan', '').replace('None', '')
        # Entferne unnötige .0 bei ganzen Zahlen
        if col in ['Fahrzeuge', 'Stopps', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppschnitt', 'Stoppkosten']:
            df_display[col] = df_display[col].apply(lambda x: x.rstrip('0').rstrip('.') if '.' in str(x) and x not in ['', 'nan', 'None'] else x)
    
    # Column Config
    column_config = {
        "Datum": st.column_config.TextColumn("Datum", help="Format: TT.MM.JJJJ"),
        "Standort": st.column_config.TextColumn("Standort"),
        "Disponent": st.column_config.TextColumn("Disponent"),
        "Fahrzeuge": st.column_config.TextColumn("Fahrzeuge"),
        "Stopps": st.column_config.TextColumn("Stopps"),
        "Unverplante Stopps": st.column_config.TextColumn("Unverplante Stopps"),
        "Kosten Fuhrpark": st.column_config.TextColumn("Kosten Fuhrpark (€)"),
        "Stoppschnitt": st.column_config.TextColumn("Stoppschnitt", disabled=True),
        "Stoppkosten": st.column_config.TextColumn("Stoppkosten", disabled=True),
    }
    
    # Spaltenreihenfolge mit Datum als erste Spalte
    column_order = ['Datum', 'Standort', 'Disponent', 'Fahrzeuge', 'Stopps', 'Unverplante Stopps', 'Kosten Fuhrpark', 'Stoppschnitt', 'Stoppkosten']
    
    edited_df = st.data_editor(
        df_display,
        num_rows="dynamic",
        column_config=column_config,
        column_order=column_order,
        hide_index=True,
        use_container_width=True
    )
    
    # Berechne Stoppschnitt und Stoppkosten automatisch
    for idx in edited_df.index:
        try:
            fahrzeuge = pd.to_numeric(str(edited_df.at[idx, 'Fahrzeuge']).replace(',', '.'), errors='coerce')
            stopps = pd.to_numeric(str(edited_df.at[idx, 'Stopps']).replace(',', '.'), errors='coerce')
            kosten = pd.to_numeric(str(edited_df.at[idx, 'Kosten Fuhrpark']).replace(',', '.'), errors='coerce')
            
            # Stoppschnitt berechnen oder löschen
            if pd.notna(fahrzeuge) and pd.notna(stopps) and fahrzeuge > 0:
                edited_df.at[idx, 'Stoppschnitt'] = str(round(stopps / fahrzeuge, 1)).replace('.', ',')
            else:
                edited_df.at[idx, 'Stoppschnitt'] = ''
            
            # Stoppkosten berechnen oder löschen
            if pd.notna(kosten) and pd.notna(stopps) and stopps > 0:
                edited_df.at[idx, 'Stoppkosten'] = str(round(kosten / stopps, 2)).replace('.', ',')
            else:
                edited_df.at[idx, 'Stoppkosten'] = ''
        except:
            pass
    
    # Validierung
    errors, warnings = validate_data(edited_df)
    if errors:
        st.error("🚨 Validierungsfehler:")
        for error in errors:
            st.error(f"❌ {error}")
    
    # Buttons
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        csv = edited_df.to_csv(index=False, sep=';').encode('utf-8')
        st.download_button("📥 CSV Export", csv, f"KPI_{selected_month}.csv")
    
    with col2:
        excel = export_to_excel(edited_df, selected_month)
        st.download_button("📊 Excel Export", excel, f"KPI_{selected_month}.xlsx")
    
    with col3:
        if st.button("💾 Speichern", type="primary", disabled=len(errors) > 0):
            save_month_data(selected_month, edited_df)
            st.success("✅ Gespeichert!")
            st.rerun()

elif page == "📊 Daily Report":
    st.header("📊 Daily Report")
    
    if not df.empty:
        df['Datum'] = pd.to_datetime(df['Datum'], dayfirst=True, errors='coerce')
        df = clean_kpi_columns(df)
        
        # Filtere Zeilen mit Daten
        df_with_data = df[(df['Stopps'].notna()) & (df['Stopps'] > 0)]
        
        if not df_with_data.empty:
            latest_date = df_with_data['Datum'].max()
            df_latest = df_with_data[df_with_data['Datum'] == latest_date]
            
            st.subheader(f"📅 {latest_date.strftime('%d.%m.%Y')}")
            
            # KPIs mit deutscher Formatierung
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Ø Stoppkosten", f"{format_number_de(df_latest['Stoppkosten'].mean(), 2)} €")
            col2.metric("Gesamt Stopps", format_number_de(int(df_latest['Stopps'].sum()), 0))
            col3.metric("Fahrzeuge", format_number_de(int(df_latest['Fahrzeuge'].sum()), 0))
            col4.metric("Ø Stoppschnitt", format_number_de(df_latest['Stoppschnitt'].mean(), 1))
            
            st.markdown("---")
            
            # Top/Bottom Standorte
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**🟢 TOP 3 - Niedrigste Stoppkosten**")
                top3 = df_latest.nsmallest(3, 'Stoppkosten')[['Standort', 'Stoppkosten', 'Stopps']].copy()
                top3['Stoppkosten'] = top3['Stoppkosten'].apply(lambda x: f"{format_number_de(x, 2)} €")
                top3['Stopps'] = top3['Stopps'].apply(lambda x: format_number_de(x, 0))
                st.dataframe(top3, hide_index=True)
            
            with col2:
                st.markdown("**🔴 BOTTOM 3 - Höchste Stoppkosten**")
                bottom3 = df_latest.nlargest(3, 'Stoppkosten')[['Standort', 'Stoppkosten', 'Stopps']].copy()
                bottom3['Stoppkosten'] = bottom3['Stoppkosten'].apply(lambda x: f"{format_number_de(x, 2)} €")
                bottom3['Stopps'] = bottom3['Stopps'].apply(lambda x: format_number_de(x, 0))
                st.dataframe(bottom3, hide_index=True)
        else:
            st.info("Keine Daten mit KPI-Werten vorhanden.")
    else:
        st.info("Keine Daten vorhanden.")

elif page == "📈 Wochenvergleich":
    st.header("📈 Wochenvergleich")
    
    if not df.empty:
        weekly_data = compare_weeks(df)
        
        if not weekly_data.empty:
            st.subheader(f"KPIs pro Woche - {selected_month}")
            
            # Formatiere Tabelle für deutsche Darstellung
            display_data = weekly_data.copy()
            
            # Formatiere numerische Spalten
            display_data['Fahrzeuge'] = display_data['Fahrzeuge'].apply(lambda x: format_number_de(x, 1))
            display_data['Stopps'] = display_data['Stopps'].apply(lambda x: format_number_de(x, 1))
            display_data['Stoppschnitt'] = display_data['Stoppschnitt'].apply(lambda x: format_number_de(x, 2))
            display_data['Unverplante Stopps'] = display_data['Unverplante Stopps'].apply(lambda x: format_number_de(x, 2))
            display_data['Kosten Fuhrpark'] = display_data['Kosten Fuhrpark'].apply(lambda x: format_number_de(x, 2) + ' €' if x != '' else '')
            display_data['Stoppkosten'] = display_data['Stoppkosten'].apply(lambda x: format_number_de(x, 2) + ' €' if x != '' else '')
            
            # Formatiere Delta-Spalten
            for col in display_data.columns:
                if 'Delta%' in col:
                    display_data[col] = display_data[col].apply(lambda x: f"{x:+.1f}%".replace('.', ',') if pd.notna(x) else "")
                elif 'Delta' in col and 'Delta%' not in col:
                    # Absolute Delta-Werte
                    if 'Kosten' in col or 'Stopp' in col:
                        display_data[col] = display_data[col].apply(lambda x: ('+' if x > 0 else '') + format_number_de(x, 2) if pd.notna(x) else "")
                    else:
                        display_data[col] = display_data[col].apply(lambda x: ('+' if x > 0 else '') + format_number_de(x, 1) if pd.notna(x) else "")
            
            st.dataframe(display_data, hide_index=True, use_container_width=True)
            
            # Charts
            st.markdown("---")
            col1, col2 = st.columns(2)
            
            with col1:
                chart = alt.Chart(weekly_data).mark_line(point=True).encode(
                    x='Woche:N',
                    y='Stopps:Q',
                    tooltip=['Woche', 'Stopps']
                ).properties(title='Stopps pro Woche', height=300)
                st.altair_chart(chart, use_container_width=True)
            
            with col2:
                chart = alt.Chart(weekly_data).mark_line(point=True, color='red').encode(
                    x='Woche:N',
                    y='Stoppkosten:Q',
                    tooltip=['Woche', alt.Tooltip('Stoppkosten:Q', format='.2f')]
                ).properties(title='Stoppkosten pro Woche', height=300)
                st.altair_chart(chart, use_container_width=True)
    else:
        st.info("Keine Daten vorhanden.")

elif page == "📅 Monatsvergleich":
    st.header("📅 Monatsvergleich")
    
    if len(monate) >= 2:
        col1, col2 = st.columns(2)
        with col1:
            month1 = st.selectbox("Monat 1", monate, index=max(0, len(monate)-2))
        with col2:
            month2 = st.selectbox("Monat 2", monate, index=len(monate)-1)
        
        if st.button("🔄 Vergleichen", type="primary"):
            df1 = get_month_data(month1)
            df2 = get_month_data(month2)
            
            if not df1.empty and not df2.empty:
                df1 = clean_kpi_columns(df1)
                df2 = clean_kpi_columns(df2)
                
                comparison = pd.DataFrame({
                    'KPI': ['Fahrzeuge', 'Stopps', 'Stoppschnitt', 'Stoppkosten'],
                    month1: [
                        df1['Fahrzeuge'].sum(),
                        df1['Stopps'].sum(),
                        df1['Stoppschnitt'].mean(),
                        df1['Stoppkosten'].mean()
                    ],
                    month2: [
                        df2['Fahrzeuge'].sum(),
                        df2['Stopps'].sum(),
                        df2['Stoppschnitt'].mean(),
                        df2['Stoppkosten'].mean()
                    ]
                })
                
                comparison['Delta'] = comparison[month2] - comparison[month1]
                comparison['Delta %'] = ((comparison[month2] / comparison[month1] - 1) * 100).round(1)
                
                # Formatiere für Anzeige
                display_comp = comparison.copy()
                for col in [month1, month2, 'Delta']:
                    display_comp[col] = display_comp.apply(
                        lambda row: format_number_de(row[col], 2) if row['KPI'] in ['Stoppschnitt', 'Stoppkosten'] else format_number_de(row[col], 1),
                        axis=1
                    )
                display_comp['Delta %'] = display_comp['Delta %'].apply(lambda x: f"{x:+.1f}%".replace('.', ',') if pd.notna(x) else "")
                
                st.dataframe(display_comp, hide_index=True, use_container_width=True)
    else:
        st.info("Mindestens 2 Monate benötigt.")

elif page == "📉 Verlauf (KPIs)":
    st.header("📉 KPI Verlauf")
    
    if not df.empty:
        df['Datum'] = pd.to_datetime(df['Datum'], dayfirst=True, errors='coerce')
        df = clean_kpi_columns(df)
        
        standorte = sorted(df['Standort'].dropna().unique())
        selected = st.multiselect("Standorte", standorte, default=standorte[:1] if standorte else [])
        
        kpi = st.selectbox("KPI", ['Stoppkosten', 'Stopps', 'Fahrzeuge', 'Stoppschnitt'])
        
        if selected:
            chart_data = df[df['Standort'].isin(selected)].sort_values('Datum')
            
            chart = alt.Chart(chart_data).mark_line(point=True).encode(
                x=alt.X('Datum:T', title='Datum'),
                y=alt.Y(f'{kpi}:Q', title=kpi),
                color='Standort:N',
                tooltip=['Datum:T', 'Standort', f'{kpi}:Q']
            ).properties(height=400)
            
            st.altair_chart(chart, use_container_width=True)
    else:
        st.info("Keine Daten vorhanden.")
